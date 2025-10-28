"""
ADVANCED INVENTORY TRACKER WITH DATABASE
Professional-grade inventory management system with past transaction entry
"""

import sqlite3
import pandas as pd
import os
import shutil
import traceback
import json
from datetime import datetime, date, timedelta
from collections import defaultdict

# Try to import rich for beautiful interface
try:
    from rich.console import Console
    from rich.table import Table
    from rich.prompt import Prompt, Confirm
    from rich.panel import Panel
    from rich import box
    RICH_AVAILABLE = True
    console = Console()
except ImportError:
    RICH_AVAILABLE = False
    print("Installing rich library for better interface...")
    os.system("pip install rich matplotlib")
    print("Please restart the program.")
    exit()

# Try to import matplotlib for charts
try:
    import matplotlib.pyplot as plt
    import matplotlib
    matplotlib.use('Agg')  # Non-interactive backend
    CHARTS_AVAILABLE = True
except ImportError:
    CHARTS_AVAILABLE = False

# Configuration
DB_NAME = 'inventory.db'
CONFIG_FILE = 'config.json'
ERROR_LOG = 'error_log.txt'
USER_FILE = 'user_config.txt'

# Default configuration
DEFAULT_CONFIG = {
    "low_stock_threshold": 10,
    "backups_to_keep": 30,
    "default_currency": "₹",
    "date_format": "%Y-%m-%d",
    "enable_charts": True,
    "enable_profit_tracking": True
}

def load_config():
    """Load configuration from file or create default."""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r') as f:
                config = json.load(f)
                # Merge with defaults for any missing keys
                for key, value in DEFAULT_CONFIG.items():
                    if key not in config:
                        config[key] = value
                return config
        except Exception as e:
            log_error(f"Error loading config: {str(e)}")
            return DEFAULT_CONFIG.copy()
    else:
        save_config(DEFAULT_CONFIG)
        return DEFAULT_CONFIG.copy()

def save_config(config):
    """Save configuration to file."""
    try:
        with open(CONFIG_FILE, 'w') as f:
            json.dump(config, f, indent=4)
    except Exception as e:
        log_error(f"Error saving config: {str(e)}")

# Load configuration
CONFIG = load_config()

def log_error(error_msg):
    """Logs errors to a file for debugging."""
    with open(ERROR_LOG, 'a') as f:
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        f.write(f"\n[{timestamp}] {error_msg}\n")
        f.write(traceback.format_exc())

def get_user_name():
    """Gets username for audit trail."""
    if not os.path.exists(USER_FILE):
        if RICH_AVAILABLE:
            name = Prompt.ask("Enter your name (for record-keeping)", default="Admin")
        else:
            name = input("Enter your name (for record-keeping): ").strip() or "Admin"
        with open(USER_FILE, 'w') as f:
            f.write(name)
        return name
    else:
        with open(USER_FILE, 'r') as f:
            return f.read().strip()

def normalize_dimension(dimension):
    """Standardizes dimension format to prevent typos creating duplicates."""
    dimension = dimension.strip().lower()
    dimension = dimension.replace('*', 'x').replace('X', 'x').replace(' ', '')
    return dimension

def init_database():
    """Initialize SQLite database with required tables."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    
    # Main transactions table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            time TEXT NOT NULL,
            user TEXT NOT NULL,
            dimension TEXT NOT NULL,
            action TEXT NOT NULL,
            amount_kg REAL NOT NULL,
            current_stock_kg REAL NOT NULL,
            cost_per_kg REAL DEFAULT 0,
            sell_per_kg REAL DEFAULT 0,
            notes TEXT DEFAULT ''
        )
    ''')
    
    # Create index for faster queries
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_dimension 
        ON transactions(dimension)
    ''')
    
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_date 
        ON transactions(date)
    ''')
    
    conn.commit()
    conn.close()
    
    # Create backup
    create_backup()

def create_backup():
    """Creates timestamped backup of database."""
    if os.path.exists(DB_NAME):
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f'backup_db_{timestamp}.db'
        shutil.copy(DB_NAME, backup_name)
        cleanup_old_backups()

def cleanup_old_backups():
    """Keeps only the most recent backups to save space."""
    try:
        backups = sorted([f for f in os.listdir('.') if f.startswith('backup_db_')])
        keep = CONFIG['backups_to_keep']
        if len(backups) > keep:
            for old_backup in backups[:-keep]:
                os.remove(old_backup)
    except Exception as e:
        log_error(f"Error cleaning up backups: {str(e)}")

def get_all_dimensions():
    """Get list of all existing dimensions for autocomplete."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('SELECT DISTINCT dimension FROM transactions ORDER BY dimension')
    dimensions = [row[0] for row in cursor.fetchall()]
    conn.close()
    return dimensions

def autocomplete_dimension(partial):
    """Suggest dimensions based on partial input."""
    dimensions = get_all_dimensions()
    matches = [d for d in dimensions if d.startswith(partial.lower())]
    return matches

def get_dimension_with_autocomplete(prompt_text):
    """Get dimension input with autocomplete suggestions."""
    dimensions = get_all_dimensions()
    
    if RICH_AVAILABLE:
        console.print(f"\n[cyan]{prompt_text}[/cyan]")
        if dimensions:
            console.print("[dim]Existing dimensions: " + ", ".join(dimensions[:10]) + 
                         ("..." if len(dimensions) > 10 else "") + "[/dim]")
        dimension = Prompt.ask("Dimension")
    else:
        print(f"\n{prompt_text}")
        if dimensions:
            print(f"Existing: {', '.join(dimensions[:10])}")
        dimension = input("Dimension: ")
    
    dimension = normalize_dimension(dimension)
    
    # Show suggestions if partial match
    if dimension and dimension not in dimensions:
        matches = autocomplete_dimension(dimension)
        if matches:
            if RICH_AVAILABLE:
                console.print(f"[yellow]Did you mean: {', '.join(matches)}?[/yellow]")
            else:
                print(f"Did you mean: {', '.join(matches)}?")
    
    return dimension

def get_current_stock(dimension):
    """Get current stock for a dimension."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT current_stock_kg 
        FROM transactions 
        WHERE dimension = ? 
        ORDER BY id DESC 
        LIMIT 1
    ''', (dimension,))
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else 0

def add_transaction(dimension, action, amount_kg, current_stock_kg, 
                   cost_per_kg=0, sell_per_kg=0, notes='', custom_date=None):
    """Add a transaction to the database."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    
    # Use custom date if provided, otherwise use today
    trans_date = custom_date if custom_date else date.today().strftime(CONFIG['date_format'])
    
    cursor.execute('''
        INSERT INTO transactions 
        (date, time, user, dimension, action, amount_kg, current_stock_kg, 
         cost_per_kg, sell_per_kg, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        trans_date,
        datetime.now().strftime('%H:%M:%S'),
        get_user_name(),
        dimension,
        action,
        amount_kg,
        current_stock_kg,
        cost_per_kg,
        sell_per_kg,
        notes
    ))
    
    conn.commit()
    conn.close()

def add_stock():
    """Records a new stock arrival with cost tracking."""
    if RICH_AVAILABLE:
        console.print("\n[bold cyan]📦 Add New Stock[/bold cyan]", style="cyan")
    else:
        print("\n--- Add New Stock ---")
    
    dimension = get_dimension_with_autocomplete("Enter bag dimension (e.g., 10x16)")
    
    if not dimension:
        console.print("[red]❌ Error: Dimension cannot be empty.[/red]")
        return
    
    try:
        if RICH_AVAILABLE:
            amount = float(Prompt.ask(f"Amount of '{dimension}' received (kg)"))
        else:
            amount = float(input(f"Amount of '{dimension}' received (kg): "))
        
        if amount <= 0:
            console.print("[red]❌ Error: Amount must be positive.[/red]")
            return
    except ValueError:
        console.print("[red]❌ Error: Invalid amount.[/red]")
        return
    
    # Ask for cost if profit tracking enabled
    cost_per_kg = 0
    if CONFIG['enable_profit_tracking']:
        try:
            if RICH_AVAILABLE:
                cost_input = Prompt.ask(f"Cost per kg {CONFIG['default_currency']}", default="0")
            else:
                cost_input = input(f"Cost per kg {CONFIG['default_currency']} (press Enter to skip): ") or "0"
            cost_per_kg = float(cost_input)
        except ValueError:
            cost_per_kg = 0
    
    current_stock = get_current_stock(dimension)
    new_stock = current_stock + amount
    
    # Confirmation
    if RICH_AVAILABLE:
        table = Table(show_header=False, box=box.ROUNDED)
        table.add_row("Dimension:", f"[cyan]{dimension}[/cyan]")
        table.add_row("Adding:", f"[green]{amount} kg[/green]")
        table.add_row("Current stock:", f"{current_stock} kg")
        table.add_row("New stock:", f"[bold green]{new_stock} kg[/bold green]")
        if cost_per_kg > 0:
            table.add_row("Cost:", f"{CONFIG['default_currency']}{cost_per_kg}/kg (Total: {CONFIG['default_currency']}{cost_per_kg * amount})")
        console.print(table)
        
        if not Confirm.ask("Is this correct?", default=True):
            console.print("[yellow]❌ Operation cancelled.[/yellow]")
            return
    else:
        print(f"\nDimension: {dimension}")
        print(f"Adding: {amount} kg")
        print(f"Current: {current_stock} kg → New: {new_stock} kg")
        if input("Correct? (yes/no): ").lower() not in ['yes', 'y']:
            return
    
    create_backup()
    add_transaction(dimension, 'Stock Added', amount, new_stock, cost_per_kg=cost_per_kg)
    
    if RICH_AVAILABLE:
        console.print(f"\n[bold green]✅ Success![/bold green] Stock for '{dimension}' is now {new_stock:.2f} kg")
    else:
        print(f"\n✅ Success! Stock for '{dimension}' is now {new_stock:.2f} kg")

def record_sale():
    """Records a sale with profit tracking."""
    if RICH_AVAILABLE:
        console.print("\n[bold cyan]💰 Record a Sale[/bold cyan]")
    else:
        print("\n--- Record a Sale ---")
    
    dimension = get_dimension_with_autocomplete("Enter bag dimension sold")
    
    if not dimension:
        console.print("[red]❌ Error: Dimension cannot be empty.[/red]")
        return
    
    current_stock = get_current_stock(dimension)
    
    if current_stock == 0:
        console.print(f"[red]❌ ERROR: No stock available for '{dimension}'![/red]")
        return
    
    if RICH_AVAILABLE:
        console.print(f"[cyan]Current stock for '{dimension}': {current_stock} kg[/cyan]")
    else:
        print(f"Current stock: {current_stock} kg")
    
    try:
        if RICH_AVAILABLE:
            amount = float(Prompt.ask("Amount sold (kg)"))
        else:
            amount = float(input("Amount sold (kg): "))
        
        if amount <= 0:
            console.print("[red]❌ Error: Amount must be positive.[/red]")
            return
    except ValueError:
        console.print("[red]❌ Error: Invalid amount.[/red]")
        return
    
    if amount > current_stock:
        console.print(f"[red]❌ ERROR: Cannot sell {amount} kg![/red]")
        console.print(f"[yellow]Only {current_stock} kg available (Short by: {amount - current_stock} kg)[/yellow]")
        return
    
    # Ask for selling price if profit tracking enabled
    sell_per_kg = 0
    if CONFIG['enable_profit_tracking']:
        try:
            if RICH_AVAILABLE:
                sell_input = Prompt.ask(f"Selling price per kg {CONFIG['default_currency']}", default="0")
            else:
                sell_input = input(f"Selling price per kg {CONFIG['default_currency']} (press Enter to skip): ") or "0"
            sell_per_kg = float(sell_input)
        except ValueError:
            sell_per_kg = 0
    
    new_stock = current_stock - amount
    
    # Confirmation
    if RICH_AVAILABLE:
        table = Table(show_header=False, box=box.ROUNDED)
        table.add_row("Dimension:", f"[cyan]{dimension}[/cyan]")
        table.add_row("Selling:", f"[yellow]{amount} kg[/yellow]")
        table.add_row("Current stock:", f"{current_stock} kg")
        table.add_row("Remaining:", f"[bold]{new_stock} kg[/bold]")
        if sell_per_kg > 0:
            table.add_row("Revenue:", f"{CONFIG['default_currency']}{sell_per_kg}/kg (Total: {CONFIG['default_currency']}{sell_per_kg * amount})")
        console.print(table)
        
        if not Confirm.ask("Is this correct?", default=True):
            console.print("[yellow]❌ Operation cancelled.[/yellow]")
            return
    else:
        print(f"\nDimension: {dimension}")
        print(f"Selling: {amount} kg")
        print(f"Current: {current_stock} kg → Remaining: {new_stock} kg")
        if input("Correct? (yes/no): ").lower() not in ['yes', 'y']:
            return
    
    create_backup()
    add_transaction(dimension, 'Sale', -amount, new_stock, sell_per_kg=sell_per_kg)
    
    if RICH_AVAILABLE:
        console.print(f"\n[bold green]✅ Success![/bold green] Stock for '{dimension}' is now {new_stock:.2f} kg")
        
        # Low stock warning
        if new_stock < CONFIG['low_stock_threshold'] and new_stock > 0:
            console.print(f"[yellow]⚠️ WARNING: Low stock! Only {new_stock:.2f} kg remaining.[/yellow]")
    else:
        print(f"\n✅ Success! Stock: {new_stock:.2f} kg")

def add_past_transaction_manual():
    """Manually enter a past transaction one at a time - user friendly!"""
    if RICH_AVAILABLE:
        console.print("\n[bold cyan]⏰ Add Past Transaction[/bold cyan]")
        console.print("[dim]Enter historical transactions one by one[/dim]\n")
    else:
        print("\n--- Add Past Transaction ---")
    
    # Step 1: Select transaction type
    if RICH_AVAILABLE:
        console.print("[bold]What type of transaction?[/bold]")
        console.print("  [green]1.[/green] Stock Added (received inventory)")
        console.print("  [yellow]2.[/yellow] Sale (sold to customer)")
        console.print("  [cyan]3.[/cyan] Adjustment (correction)")
        
        trans_type = Prompt.ask("Choose transaction type", choices=['1','2','3'])
    else:
        print("\nTransaction Type:")
        print("1. Stock Added")
        print("2. Sale")
        print("3. Adjustment")
        trans_type = input("Choice [1]: ").strip() or '1'
    
    # Map choice to action
    if trans_type == '1':
        action = 'Stock Added'
        action_color = 'green'
        action_symbol = '+'
    elif trans_type == '2':
        action = 'Sale'
        action_color = 'yellow'
        action_symbol = '-'
    else:
        action = 'Adjustment'
        action_color = 'cyan'
        action_symbol = '±'
    
    # Step 2: Get date
    if RICH_AVAILABLE:
        console.print(f"\n[bold]When did this {action.lower()} occur?[/bold]")
        custom_date = Prompt.ask("Enter date (YYYY-MM-DD)", 
                                default=date.today().strftime('%Y-%m-%d'))
    else:
        print(f"\nDate of {action}:")
        custom_date = input(f"Enter date (YYYY-MM-DD) [{date.today().strftime('%Y-%m-%d')}]: ") or date.today().strftime('%Y-%m-%d')
    
    # Validate date
    try:
        date_obj = datetime.strptime(custom_date, '%Y-%m-%d')
        if date_obj > datetime.now():
            console.print("[red]❌ Cannot enter future dates![/red]")
            return
    except ValueError:
        console.print("[red]❌ Invalid date format. Use YYYY-MM-DD[/red]")
        return
    
    # Step 3: Get dimension
    dimension = get_dimension_with_autocomplete(f"Enter bag dimension for this {action.lower()}")
    
    if not dimension:
        console.print("[red]❌ Dimension cannot be empty.[/red]")
        return
    
    # Step 4: Get amount
    try:
        if RICH_AVAILABLE:
            if action == 'Sale':
                amount = float(Prompt.ask(f"Amount sold (kg)", default="0"))
            elif action == 'Stock Added':
                amount = float(Prompt.ask(f"Amount received (kg)", default="0"))
            else:
                amount = float(Prompt.ask(f"Adjustment amount (use + or - for direction)", default="0"))
        else:
            amount = float(input(f"Amount (kg): ") or "0")
        
        if amount == 0:
            console.print("[red]❌ Amount cannot be zero.[/red]")
            return
    except ValueError:
        console.print("[red]❌ Invalid amount.[/red]")
        return
    
    # Step 5: Get pricing info
    cost_per_kg = 0
    sell_per_kg = 0
    
    if CONFIG['enable_profit_tracking']:
        if action == 'Stock Added':
            try:
                if RICH_AVAILABLE:
                    cost_input = Prompt.ask(f"Cost per kg {CONFIG['default_currency']} (optional)", default="0")
                else:
                    cost_input = input(f"Cost per kg {CONFIG['default_currency']} [0]: ") or "0"
                cost_per_kg = float(cost_input)
            except ValueError:
                cost_per_kg = 0
        
        elif action == 'Sale':
            try:
                if RICH_AVAILABLE:
                    sell_input = Prompt.ask(f"Selling price per kg {CONFIG['default_currency']} (optional)", default="0")
                else:
                    sell_input = input(f"Selling price per kg {CONFIG['default_currency']} [0]: ") or "0"
                sell_per_kg = float(sell_input)
            except ValueError:
                sell_per_kg = 0
    
    # Step 6: Get notes
    if RICH_AVAILABLE:
        notes = Prompt.ask("Notes (optional)", default="")
    else:
        notes = input("Notes (optional): ").strip()
    
    # Step 7: Calculate new stock
    current_stock = get_current_stock(dimension)
    
    if action == 'Stock Added':
        new_stock = current_stock + abs(amount)
        amount_signed = abs(amount)
    elif action == 'Sale':
        amount_signed = -abs(amount)
        new_stock = current_stock + amount_signed
        if new_stock < 0:
            console.print(f"[yellow]⚠️ Warning: This will result in negative stock ({new_stock:.2f} kg)[/yellow]")
            if RICH_AVAILABLE:
                if not Confirm.ask("Continue anyway?", default=False):
                    console.print("[yellow]❌ Transaction cancelled.[/yellow]")
                    return
    else:  # Adjustment
        amount_signed = amount
        new_stock = current_stock + amount
    
    # Step 8: Show summary and confirm
    if RICH_AVAILABLE:
        console.print("\n[bold]Transaction Summary:[/bold]")
        
        table = Table(show_header=False, box=box.ROUNDED, border_style=action_color)
        table.add_row("Date:", f"[cyan]{custom_date}[/cyan]")
        table.add_row("Type:", f"[{action_color}]{action}[/{action_color}]")
        table.add_row("Dimension:", f"[cyan]{dimension}[/cyan]")
        table.add_row("Amount:", f"[{action_color}]{action_symbol}{abs(amount):.2f} kg[/{action_color}]")
        table.add_row("Current Stock:", f"{current_stock:.2f} kg")
        table.add_row("New Stock:", f"[bold]{new_stock:.2f} kg[/bold]")
        
        if cost_per_kg > 0:
            total_cost = abs(amount) * cost_per_kg
            table.add_row("Cost:", f"{CONFIG['default_currency']}{cost_per_kg}/kg (Total: {CONFIG['default_currency']}{total_cost:.2f})")
        
        if sell_per_kg > 0:
            total_revenue = abs(amount) * sell_per_kg
            table.add_row("Revenue:", f"{CONFIG['default_currency']}{sell_per_kg}/kg (Total: {CONFIG['default_currency']}{total_revenue:.2f})")
        
        if notes:
            table.add_row("Notes:", f"[dim]{notes}[/dim]")
        
        console.print(table)
        
        if not Confirm.ask("\nIs this correct?", default=True):
            console.print("[yellow]❌ Transaction cancelled.[/yellow]")
            return
    else:
        print(f"\nSummary:")
        print(f"Date: {custom_date}")
        print(f"Type: {action}")
        print(f"Dimension: {dimension}")
        print(f"Amount: {action_symbol}{abs(amount):.2f} kg")
        print(f"Stock: {current_stock:.2f} kg → {new_stock:.2f} kg")
        if input("\nConfirm? (yes/no): ").lower() not in ['yes', 'y']:
            return
    
    # Step 9: Save transaction
    create_backup()
    add_transaction(dimension, action, amount_signed, new_stock, 
                   cost_per_kg, sell_per_kg, notes, custom_date)
    
    if RICH_AVAILABLE:
        console.print(f"\n[bold green]✅ Past transaction recorded successfully![/bold green]")
        console.print(f"[dim]Date: {custom_date} | {dimension}: {new_stock:.2f} kg[/dim]")
        
        # Ask if they want to add another
        console.print()
        if Confirm.ask("Add another past transaction?", default=False):
            add_past_transaction_manual()
    else:
        print(f"\n✅ Transaction recorded for {custom_date}")

def bulk_entry_wizard():
    """Guide user through entering multiple past transactions easily."""
    if RICH_AVAILABLE:
        console.print("\n[bold cyan]📋 Bulk Entry Wizard[/bold cyan]")
        console.print("[dim]Enter multiple past transactions quickly[/dim]\n")
        console.print("[yellow]Tip: Have your records ready (dates, amounts, prices, etc.)[/yellow]\n")
        
        # Ask if user wants to include pricing
        include_pricing = False
        if CONFIG['enable_profit_tracking']:
            include_pricing = Confirm.ask("Include pricing information?", default=True)
    else:
        print("\n--- Bulk Entry Wizard ---")
        include_pricing = False
        if CONFIG['enable_profit_tracking']:
            include_pricing = input("Include pricing? (y/n) [y]: ").lower() in ['', 'y', 'yes']
    
    transactions_added = 0
    
    while True:
        if RICH_AVAILABLE:
            console.print(f"\n[bold]Transaction #{transactions_added + 1}[/bold]")
        else:
            print(f"\n--- Transaction #{transactions_added + 1} ---")
        
        # Quick entry mode
        try:
            # Get all info quickly
            if RICH_AVAILABLE:
                date_input = Prompt.ask("Date (YYYY-MM-DD)", default=date.today().strftime('%Y-%m-%d'))
            else:
                date_input = input(f"Date (YYYY-MM-DD) [{date.today().strftime('%Y-%m-%d')}]: ") or date.today().strftime('%Y-%m-%d')
            
            # Validate date
            datetime.strptime(date_input, '%Y-%m-%d')
            
            dimension = get_dimension_with_autocomplete("Dimension")
            if not dimension:
                break
            
            if RICH_AVAILABLE:
                trans_type = Prompt.ask("Type (stock/sale/adjust)", 
                                      choices=['stock', 'sale', 'adjust'])
            else:
                trans_type = input("Type (stock/sale/adjust): ").lower()
            
            if trans_type not in ['stock', 'sale', 'adjust']:
                console.print("[red]Invalid type[/red]")
                continue
            
            amount = float(input("Amount (kg): "))
            
            # Get pricing if enabled
            cost_per_kg = 0
            sell_per_kg = 0
            
            if include_pricing:
                if trans_type == 'stock':
                    try:
                        cost_input = input(f"Cost per kg {CONFIG['default_currency']} [0]: ") or "0"
                        cost_per_kg = float(cost_input)
                    except ValueError:
                        cost_per_kg = 0
                elif trans_type == 'sale':
                    try:
                        sell_input = input(f"Sell price per kg {CONFIG['default_currency']} [0]: ") or "0"
                        sell_per_kg = float(sell_input)
                    except ValueError:
                        sell_per_kg = 0
            
            # Map type
            if trans_type == 'stock':
                action = 'Stock Added'
                amount_signed = abs(amount)
            elif trans_type == 'sale':
                action = 'Sale'
                amount_signed = -abs(amount)
            else:
                action = 'Adjustment'
                amount_signed = amount
            
            # Calculate stock
            current_stock = get_current_stock(dimension)
            new_stock = current_stock + amount_signed
            
            # Quick confirm with pricing info
            summary = f"→ {date_input} | {dimension} | {action} | {amount_signed:+.2f}kg | Stock: {new_stock:.2f}kg"
            if cost_per_kg > 0:
                summary += f" | Cost: {CONFIG['default_currency']}{cost_per_kg}/kg"
            if sell_per_kg > 0:
                summary += f" | Price: {CONFIG['default_currency']}{sell_per_kg}/kg"
            
            print(summary)
            
            if RICH_AVAILABLE:
                confirm = Confirm.ask("OK?", default=True)
            else:
                confirm = input("OK? (y/n) [y]: ").lower() in ['', 'y', 'yes']
            
            if confirm:
                create_backup()
                add_transaction(dimension, action, amount_signed, new_stock, 
                              cost_per_kg=cost_per_kg, sell_per_kg=sell_per_kg, 
                              notes="Bulk entry", custom_date=date_input)
                transactions_added += 1
                console.print(f"[green]✅ Added ({transactions_added} total)[/green]")
            else:
                console.print("[yellow]⏭️ Skipped[/yellow]")
        
        except KeyboardInterrupt:
            console.print("\n[yellow]Bulk entry interrupted[/yellow]")
            break
        except Exception as e:
            console.print(f"[red]Error: {str(e)}[/red]")
            continue
        
        # Continue?
        if RICH_AVAILABLE:
            if not Confirm.ask("\nAdd another?", default=True):
                break
        else:
            if input("\nAdd another? (y/n) [y]: ").lower() not in ['', 'y', 'yes']:
                break
    
    console.print(f"\n[bold green]✅ Bulk entry complete! Added {transactions_added} transactions.[/bold green]")

def view_inventory():
    """Display current inventory with rich formatting."""
    conn = sqlite3.connect(DB_NAME)
    
    # Get current stock for each dimension
    query = '''
        SELECT dimension, current_stock_kg
        FROM transactions t1
        WHERE id = (
            SELECT MAX(id) 
            FROM transactions t2 
            WHERE t2.dimension = t1.dimension
        )
        ORDER BY dimension
    '''
    
    df = pd.read_sql_query(query, conn)
    conn.close()
    
    if df.empty:
        console.print("\n[yellow]Inventory is empty.[/yellow]")
        return
    
    if RICH_AVAILABLE:
        table = Table(title="📦 Current Inventory", box=box.DOUBLE_EDGE)
        table.add_column("Dimension", style="cyan")
        table.add_column("Stock (kg)", justify="right", style="green")
        table.add_column("Status", justify="center")
        
        total = 0
        for _, row in df.iterrows():
            stock = row['current_stock_kg']
            total += stock
            
            if stock == 0:
                status = "[red]OUT OF STOCK[/red]"
                stock_str = "[red]0.00[/red]"
            elif stock < CONFIG['low_stock_threshold']:
                status = "[yellow]⚠️ LOW[/yellow]"
                stock_str = f"[yellow]{stock:.2f}[/yellow]"
            else:
                status = "[green]✓ OK[/green]"
                stock_str = f"[green]{stock:.2f}[/green]"
            
            table.add_row(row['dimension'], stock_str, status)
        
        table.add_section()
        table.add_row("[bold]TOTAL[/bold]", f"[bold]{total:.2f}[/bold]", "")
        
        console.print(table)
    else:
        print("\n--- Current Inventory ---")
        for _, row in df.iterrows():
            print(f"{row['dimension']}: {row['current_stock_kg']:.2f} kg")

def generate_stock_chart():
    """Generate a bar chart of current stock levels."""
    if not CHARTS_AVAILABLE:
        console.print("[yellow]Chart generation not available. Install matplotlib.[/yellow]")
        return
    
    conn = sqlite3.connect(DB_NAME)
    
    query = '''
        SELECT dimension, current_stock_kg
        FROM transactions t1
        WHERE id = (
            SELECT MAX(id) 
            FROM transactions t2 
            WHERE t2.dimension = t1.dimension
        )
        ORDER BY current_stock_kg DESC
    '''
    
    df = pd.read_sql_query(query, conn)
    conn.close()
    
    if df.empty:
        console.print("[yellow]No data available for chart.[/yellow]")
        return
    
    # Create chart
    plt.figure(figsize=(12, 6))
    colors = ['red' if x == 0 else 'orange' if x < CONFIG['low_stock_threshold'] else 'green' 
              for x in df['current_stock_kg']]
    
    plt.bar(df['dimension'], df['current_stock_kg'], color=colors)
    plt.xlabel('Bag Dimension')
    plt.ylabel('Stock (kg)')
    plt.title('Current Stock Levels')
    plt.xticks(rotation=45, ha='right')
    plt.axhline(y=CONFIG['low_stock_threshold'], color='orange', linestyle='--', 
                label=f'Low Stock Threshold ({CONFIG["low_stock_threshold"]} kg)')
    plt.legend()
    plt.tight_layout()
    
    filename = f'stock_chart_{date.today().strftime("%Y%m%d")}.png'
    plt.savefig(filename, dpi=150)
    plt.close()
    
    console.print(f"[green]✅ Chart saved as: {filename}[/green]")

def stock_adjustment():
    """Manual stock adjustment for corrections."""
    if RICH_AVAILABLE:
        console.print("\n[bold yellow]⚙️ Stock Adjustment[/bold yellow]")
        console.print("[dim]Use this to correct stock levels based on physical count[/dim]")
    else:
        print("\n--- Stock Adjustment ---")
    
    dimension = get_dimension_with_autocomplete("Enter dimension to adjust")
    
    if not dimension:
        return
    
    current_stock = get_current_stock(dimension)
    
    if RICH_AVAILABLE:
        console.print(f"[cyan]Current stock in system: {current_stock} kg[/cyan]")
        actual_stock = float(Prompt.ask("Enter actual stock (from physical count)"))
    else:
        print(f"Current stock: {current_stock} kg")
        actual_stock = float(input("Actual stock (kg): "))
    
    difference = actual_stock - current_stock
    
    if difference == 0:
        console.print("[green]✓ Stock level is correct. No adjustment needed.[/green]")
        return
    
    # Confirmation
    if RICH_AVAILABLE:
        console.print(f"\n[yellow]Adjustment needed: {difference:+.2f} kg[/yellow]")
        notes = Prompt.ask("Reason for adjustment", default="Physical count correction")
        
        if not Confirm.ask("Apply this adjustment?", default=True):
            console.print("[yellow]Adjustment cancelled.[/yellow]")
            return
    else:
        print(f"\nAdjustment: {difference:+.2f} kg")
        notes = input("Reason: ") or "Physical count correction"
        if input("Apply? (yes/no): ").lower() not in ['yes', 'y']:
            return
    
    create_backup()
    add_transaction(dimension, 'Adjustment', difference, actual_stock, notes=notes)
    
    console.print(f"[green]✅ Stock adjusted. New level: {actual_stock} kg[/green]")

def view_item_history():
    """View complete transaction history for a specific dimension."""
    if RICH_AVAILABLE:
        console.print("\n[bold cyan]📜 View Item History[/bold cyan]")
    else:
        print("\n--- View Item History ---")
    
    dimension = get_dimension_with_autocomplete("Enter dimension to view history")
    
    if not dimension:
        return
    
    conn = sqlite3.connect(DB_NAME)
    query = '''
        SELECT date, time, user, action, amount_kg, current_stock_kg, 
               cost_per_kg, sell_per_kg, notes
        FROM transactions
        WHERE dimension = ?
        ORDER BY date DESC, time DESC
        LIMIT 50
    '''
    
    df = pd.read_sql_query(query, conn, params=(dimension,))
    conn.close()
    
    if df.empty:
        console.print(f"[yellow]No history found for '{dimension}'[/yellow]")
        return
    
    if RICH_AVAILABLE:
        table = Table(title=f"Transaction History for: {dimension}", box=box.DOUBLE_EDGE)
        table.add_column("Date", style="cyan")
        table.add_column("Time", style="dim")
        table.add_column("User", style="blue")
        table.add_column("Action", style="yellow")
        table.add_column("Amount (kg)", justify="right")
        table.add_column("New Stock", justify="right", style="green")
        table.add_column("Value", justify="right")
        
        for _, row in df.iterrows():
            action_color = "green" if row['action'] == "Stock Added" else "yellow" if row['action'] == "Sale" else "cyan"
            amount_str = f"{row['amount_kg']:+.2f}" if row['amount_kg'] >= 0 else f"{row['amount_kg']:.2f}"
            
            # Calculate value
            value = ""
            if row['action'] == "Stock Added" and row['cost_per_kg'] > 0:
                value = f"{CONFIG['default_currency']}{abs(row['amount_kg']) * row['cost_per_kg']:.2f}"
            elif row['action'] == "Sale" and row['sell_per_kg'] > 0:
                value = f"{CONFIG['default_currency']}{abs(row['amount_kg']) * row['sell_per_kg']:.2f}"
            
            table.add_row(
                row['date'],
                row['time'],
                row['user'],
                f"[{action_color}]{row['action']}[/{action_color}]",
                amount_str,
                f"{row['current_stock_kg']:.2f}",
                value
            )
        
        console.print(table)
        console.print(f"\n[dim]Showing last 50 transactions for {dimension} (sorted by date)[/dim]")
    else:
        print(f"\nTransaction History for: {dimension}")
        print(df.to_string(index=False))

def get_date_range():
    """Get date range from user for filtered reports."""
    if RICH_AVAILABLE:
        console.print("\n[bold cyan]📅 Select Date Range[/bold cyan]")
        console.print("1. Last 7 Days")
        console.print("2. Last 30 Days")
        console.print("3. This Month")
        console.print("4. This Year")
        console.print("5. All Time")
        console.print("6. Custom Range")
        
        choice = Prompt.ask("Choose", choices=['1','2','3','4','5','6'], default='2')
    else:
        print("\n--- Select Date Range ---")
        print("1. Last 7 Days")
        print("2. Last 30 Days")
        print("3. This Month")
        print("4. This Year")
        print("5. All Time")
        print("6. Custom Range")
        choice = input("Choice [2]: ").strip() or '2'
    
    today = date.today()
    
    if choice == '1':
        start_date = (today - timedelta(days=7)).strftime(CONFIG['date_format'])
        end_date = today.strftime(CONFIG['date_format'])
        label = "Last 7 Days"
    elif choice == '2':
        start_date = (today - timedelta(days=30)).strftime(CONFIG['date_format'])
        end_date = today.strftime(CONFIG['date_format'])
        label = "Last 30 Days"
    elif choice == '3':
        start_date = today.replace(day=1).strftime(CONFIG['date_format'])
        end_date = today.strftime(CONFIG['date_format'])
        label = "This Month"
    elif choice == '4':
        start_date = today.replace(month=1, day=1).strftime(CONFIG['date_format'])
        end_date = today.strftime(CONFIG['date_format'])
        label = "This Year"
    elif choice == '5':
        start_date = '1900-01-01'
        end_date = today.strftime(CONFIG['date_format'])
        label = "All Time"
    else:  # Custom
        if RICH_AVAILABLE:
            start_date = Prompt.ask("Start date (YYYY-MM-DD)")
            end_date = Prompt.ask("End date (YYYY-MM-DD)")
        else:
            start_date = input("Start date (YYYY-MM-DD): ")
            end_date = input("End date (YYYY-MM-DD): ")
        label = f"{start_date} to {end_date}"
    
    return start_date, end_date, label

def sales_and_profit_reports():
    """Comprehensive sales and profit reports with date filtering."""
    start_date, end_date, label = get_date_range()
    
    conn = sqlite3.connect(DB_NAME)
    
    # Sales summary
    sales_query = '''
        SELECT dimension, 
               SUM(ABS(amount_kg)) as total_sold,
               SUM(ABS(amount_kg) * sell_per_kg) as revenue
        FROM transactions
        WHERE action = 'Sale' AND date >= ? AND date <= ?
        GROUP BY dimension
        ORDER BY total_sold DESC
    '''
    
    sales_df = pd.read_sql_query(sales_query, conn, params=(start_date, end_date))
    
    # Cost summary
    cost_query = '''
        SELECT SUM(ABS(amount_kg) * cost_per_kg) as total_cost
        FROM transactions
        WHERE action = 'Stock Added' AND date >= ? AND date <= ? AND cost_per_kg > 0
    '''
    
    cost_result = pd.read_sql_query(cost_query, conn, params=(start_date, end_date))
    total_cost = cost_result['total_cost'].iloc[0] if not cost_result.empty else 0
    
    # Calculate sales velocity for the period
    velocity_query = '''
        SELECT dimension, SUM(ABS(amount_kg)) as total_sold
        FROM transactions
        WHERE action = 'Sale' AND date >= ? AND date <= ?
        GROUP BY dimension
    '''
    
    velocity_df = pd.read_sql_query(velocity_query, conn, params=(start_date, end_date))
    
    conn.close()
    
    if RICH_AVAILABLE:
        console.print(f"\n[bold cyan]📊 Sales & Profit Report: {label}[/bold cyan]")
        
        if not sales_df.empty:
            # Sales by dimension
            table = Table(title="Sales by Dimension", box=box.ROUNDED)
            table.add_column("Dimension", style="cyan")
            table.add_column("Sold (kg)", justify="right", style="yellow")
            table.add_column("Revenue", justify="right", style="green")
            
            total_sold = 0
            total_revenue = 0
            
            for _, row in sales_df.iterrows():
                total_sold += row['total_sold']
                total_revenue += row['revenue'] if row['revenue'] else 0
                
                table.add_row(
                    row['dimension'],
                    f"{row['total_sold']:.2f}",
                    f"{CONFIG['default_currency']}{row['revenue']:.2f}" if row['revenue'] else "-"
                )
            
            table.add_section()
            table.add_row("[bold]TOTAL[/bold]", f"[bold]{total_sold:.2f}[/bold]", 
                         f"[bold]{CONFIG['default_currency']}{total_revenue:.2f}[/bold]")
            
            console.print(table)
            
            # Profit summary
            if CONFIG['enable_profit_tracking'] and (total_cost > 0 or total_revenue > 0):
                profit = total_revenue - total_cost
                margin = (profit / total_revenue * 100) if total_revenue > 0 else 0
                
                profit_table = Table(title="Profit Summary", box=box.DOUBLE_EDGE)
                profit_table.add_column("Metric", style="cyan")
                profit_table.add_column("Amount", justify="right", style="green")
                
                profit_table.add_row("Total Cost", f"{CONFIG['default_currency']}{total_cost:.2f}")
                profit_table.add_row("Total Revenue", f"{CONFIG['default_currency']}{total_revenue:.2f}")
                profit_table.add_section()
                
                profit_color = "green" if profit >= 0 else "red"
                profit_table.add_row(
                    "[bold]Net Profit[/bold]", 
                    f"[bold {profit_color}]{CONFIG['default_currency']}{profit:.2f}[/bold {profit_color}]"
                )
                profit_table.add_row("Profit Margin", f"{margin:.1f}%")
                
                console.print(profit_table)
            
            # Sales velocity & forecast
            if not velocity_df.empty:
                # Calculate number of days in period
                start_dt = datetime.strptime(start_date, CONFIG['date_format'])
                end_dt = datetime.strptime(end_date, CONFIG['date_format'])
                days_in_period = (end_dt - start_dt).days + 1
                
                velocity_table = Table(title="Sales Velocity & Forecast", box=box.ROUNDED)
                velocity_table.add_column("Dimension", style="cyan")
                velocity_table.add_column("Avg/Day", justify="right")
                velocity_table.add_column("Current Stock", justify="right")
                velocity_table.add_column("Days Left", justify="right")
                
                for _, row in velocity_df.iterrows():
                    dimension = row['dimension']
                    total_sold = row['total_sold']
                    avg_per_day = total_sold / days_in_period if days_in_period > 0 else 0
                    current_stock = get_current_stock(dimension)
                    
                    if avg_per_day > 0:
                        days_remaining = current_stock / avg_per_day
                        
                        if days_remaining < 7:
                            forecast = f"[red]{days_remaining:.0f}[/red]"
                        elif days_remaining < 14:
                            forecast = f"[yellow]{days_remaining:.0f}[/yellow]"
                        else:
                            forecast = f"[green]{days_remaining:.0f}[/green]"
                    else:
                        forecast = "[dim]-[/dim]"
                    
                    velocity_table.add_row(
                        dimension,
                        f"{avg_per_day:.2f}",
                        f"{current_stock:.2f}",
                        forecast
                    )
                
                console.print(velocity_table)
        else:
            console.print(f"[yellow]No sales data for {label}[/yellow]")
    else:
        print(f"\n--- Sales & Profit Report: {label} ---")
        if not sales_df.empty:
            print(sales_df.to_string(index=False))

def undo_last_transaction():
    """Undo the last transaction with confirmation."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    
    # Get last transaction
    cursor.execute('SELECT * FROM transactions ORDER BY id DESC LIMIT 1')
    last = cursor.fetchone()
    
    if not last:
        console.print("\n[yellow]No transactions to undo.[/yellow]")
        conn.close()
        return
    
    # Parse transaction details
    trans_id, trans_date, trans_time, trans_user, dimension, action, amount_kg, \
    current_stock, cost_per_kg, sell_per_kg, notes = last
    
    if RICH_AVAILABLE:
        console.print("\n[bold red]⚠️ UNDO LAST TRANSACTION[/bold red]")
        
        table = Table(box=box.HEAVY_EDGE, border_style="red")
        table.add_column("Field", style="yellow")
        table.add_column("Value", style="white")
        
        table.add_row("Date", f"{trans_date} {trans_time}")
        table.add_row("User", trans_user)
        table.add_row("Dimension", dimension)
        table.add_row("Action", action)
        table.add_row("Amount", f"{amount_kg:+.2f} kg")
        table.add_row("Stock After", f"{current_stock:.2f} kg")
        if notes:
            table.add_row("Notes", notes)
        
        console.print(table)
        console.print("\n[red bold]WARNING: This action cannot be reversed![/red bold]")
        
        if not Confirm.ask("Are you ABSOLUTELY SURE you want to delete this?", default=False):
            console.print("[yellow]Undo cancelled.[/yellow]")
            conn.close()
            return
    else:
        print("\n--- UNDO LAST TRANSACTION ---")
        print(f"Date: {trans_date} {trans_time}")
        print(f"User: {trans_user}")
        print(f"Dimension: {dimension}")
        print(f"Action: {action}")
        print(f"Amount: {amount_kg:+.2f} kg")
        print(f"Stock After: {current_stock:.2f} kg")
        
        confirm = input("\nAre you SURE? Type 'DELETE' to confirm: ")
        if confirm != 'DELETE':
            print("Undo cancelled.")
            conn.close()
            return
    
    # Create backup before deletion
    create_backup()
    
    # Delete the transaction
    cursor.execute('DELETE FROM transactions WHERE id = ?', (trans_id,))
    conn.commit()
    conn.close()
    
    console.print("[green]✅ Transaction deleted successfully.[/green]")
    console.print(f"[dim]Note: Stock level for '{dimension}' should be verified.[/dim]")

def export_to_excel():
    """Export database to Excel with multiple professional sheets."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        console.print("[yellow]Excel export requires openpyxl. Installing...[/yellow]")
        os.system("pip install openpyxl")
        return
    
    filename = f'inventory_report_{date.today().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    conn = sqlite3.connect(DB_NAME)
    
    # Sheet 1: Current Stock Summary
    stock_query = '''
        SELECT dimension, current_stock_kg as stock_kg
        FROM transactions t1
        WHERE id = (
            SELECT MAX(id) 
            FROM transactions t2 
            WHERE t2.dimension = t1.dimension
        )
        ORDER BY dimension
    '''
    stock_df = pd.read_sql_query(stock_query, conn)
    stock_df.columns = ['Dimension', 'Stock (kg)']
    
    # Sheet 2: Transaction History
    trans_df = pd.read_sql_query(
        'SELECT date, time, user, dimension, action, amount_kg, current_stock_kg FROM transactions ORDER BY date DESC, time DESC', 
        conn
    )
    trans_df.columns = ['Date', 'Time', 'User', 'Dimension', 'Action', 'Amount (kg)', 'Stock After (kg)']
    
    # Sheet 3: Sales by Dimension
    sales_query = '''
        SELECT dimension, 
               COUNT(*) as num_sales,
               SUM(ABS(amount_kg)) as total_sold,
               AVG(ABS(amount_kg)) as avg_sale,
               SUM(ABS(amount_kg) * sell_per_kg) as revenue
        FROM transactions
        WHERE action = 'Sale'
        GROUP BY dimension
        ORDER BY total_sold DESC
    '''
    sales_df = pd.read_sql_query(sales_query, conn)
    sales_df.columns = ['Dimension', 'Num Sales', 'Total Sold (kg)', 'Avg Sale (kg)', f'Revenue ({CONFIG["default_currency"]})']
    
    # Sheet 4: Profit Analysis (if enabled)
    profit_data = None
    if CONFIG['enable_profit_tracking']:
        # Cost
        cost_query = '''
            SELECT dimension,
                   SUM(ABS(amount_kg)) as qty_purchased,
                   SUM(ABS(amount_kg) * cost_per_kg) as total_cost
            FROM transactions
            WHERE action = 'Stock Added' AND cost_per_kg > 0
            GROUP BY dimension
        '''
        cost_df = pd.read_sql_query(cost_query, conn)
        
        # Revenue
        revenue_query = '''
            SELECT dimension,
                   SUM(ABS(amount_kg)) as qty_sold,
                   SUM(ABS(amount_kg) * sell_per_kg) as total_revenue
            FROM transactions
            WHERE action = 'Sale' AND sell_per_kg > 0
            GROUP BY dimension
        '''
        revenue_df = pd.read_sql_query(revenue_query, conn)
        
        # Merge
        if not cost_df.empty and not revenue_df.empty:
            profit_df = pd.merge(cost_df, revenue_df, on='dimension', how='outer').fillna(0)
            profit_df['profit'] = profit_df['total_revenue'] - profit_df['total_cost']
            profit_df['margin_%'] = (profit_df['profit'] / profit_df['total_revenue'] * 100).fillna(0)
            profit_df.columns = ['Dimension', 'Qty Purchased', f'Total Cost ({CONFIG["default_currency"]})', 
                                'Qty Sold', f'Revenue ({CONFIG["default_currency"]})', 
                                f'Profit ({CONFIG["default_currency"]})', 'Margin (%)']
            profit_data = profit_df
    
    conn.close()
    
    # Write to Excel with multiple sheets
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        stock_df.to_excel(writer, sheet_name='Current Stock', index=False)
        trans_df.to_excel(writer, sheet_name='Transaction History', index=False)
        sales_df.to_excel(writer, sheet_name='Sales Summary', index=False)
        if profit_data is not None:
            profit_data.to_excel(writer, sheet_name='Profit Analysis', index=False)
    
    # Format the Excel file
    wb = load_workbook(filename)
    
    # Format Current Stock sheet
    ws = wb['Current Stock']
    
    # Header formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Color code stock levels
    low_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    good_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    for row in range(2, ws.max_row + 1):
        stock_value = ws.cell(row, 2).value
        if stock_value is not None:
            if stock_value == 0:
                ws.cell(row, 2).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                ws.cell(row, 2).font = Font(color='FFFFFF', bold=True)
            elif stock_value < CONFIG['low_stock_threshold']:
                ws.cell(row, 2).fill = low_fill
            else:
                ws.cell(row, 2).fill = good_fill
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)
    
    console.print(f"\n[green]✅ Professional Excel report created: {filename}[/green]")
    console.print("[dim]Contains: Current Stock, Transaction History, Sales Summary, Profit Analysis[/dim]")

def edit_settings():
    """Edit configuration settings."""
    if RICH_AVAILABLE:
        console.print("\n[bold cyan]⚙️ Settings[/bold cyan]")
        
        table = Table(box=box.ROUNDED)
        table.add_column("Setting", style="cyan")
        table.add_column("Current Value", style="green")
        
        for key, value in CONFIG.items():
            table.add_row(key, str(value))
        
        console.print(table)
        
        if Confirm.ask("\nEdit settings?", default=False):
            CONFIG['low_stock_threshold'] = float(Prompt.ask(
                "Low stock threshold (kg)", 
                default=str(CONFIG['low_stock_threshold'])
            ))
            CONFIG['backups_to_keep'] = int(Prompt.ask(
                "Number of backups to keep", 
                default=str(CONFIG['backups_to_keep'])
            ))
            
            save_config(CONFIG)
            console.print("[green]✅ Settings saved![/green]")

def main_menu():
    """Display main menu and handle user choice."""
    user = get_user_name()
    
    while True:
        if RICH_AVAILABLE:
            console.print("\n" + "="*60)
            console.print(Panel.fit(
                "[bold cyan]BIODEGRADABLE BAGS INVENTORY SYSTEM[/bold cyan]\n"
                f"[dim]User: {user} | Date: {date.today().strftime('%d %B %Y')}[/dim]",
                border_style="cyan"
            ))
            console.print("="*60)
            
            console.print("\n[bold]📦 Stock Operations:[/bold]")
            console.print("  [cyan]1.[/cyan] Add New Stock (Today)")
            console.print("  [cyan]2.[/cyan] Record Sale (Today)")
            
            console.print("\n[bold]⏰ Past Transactions:[/bold]")
            console.print("  [cyan]3.[/cyan] Add Past Transaction (Manual Entry)")
            console.print("  [cyan]4.[/cyan] Bulk Entry Wizard (Multiple Past Transactions)")
            
            console.print("\n[bold]📊 Reports & Views:[/bold]")
            console.print("  [cyan]5.[/cyan] View Current Inventory")
            console.print("  [cyan]6.[/cyan] View Item History")
            console.print("  [cyan]7.[/cyan] Sales & Profit Reports")
            console.print("  [cyan]8.[/cyan] Generate Stock Chart")
            
            console.print("\n[bold]🔧 Management:[/bold]")
            console.print("  [cyan]9.[/cyan] Stock Adjustment (Physical Count)")
            console.print("  [cyan]10.[/cyan] Undo Last Transaction")
            console.print("  [cyan]11.[/cyan] Export to Excel")
            console.print("  [cyan]12.[/cyan] Settings")
            console.print("  [cyan]0.[/cyan] Exit")
            
            choice = Prompt.ask("\nEnter your choice", 
                              choices=['1','2','3','4','5','6','7','8','9','10','11','12','0'])
        else:
            print("\n" + "="*50)
            print("INVENTORY SYSTEM - MAIN MENU")
            print("="*50)
            print("\nStock Operations:")
            print("1. Add New Stock (Today)")
            print("2. Record Sale (Today)")
            print("\nPast Transactions:")
            print("3. Add Past Transaction")
            print("4. Bulk Entry Wizard")
            print("\nReports:")
            print("5. View Current Inventory")
            print("6. View Item History")
            print("7. Sales & Profit Reports")
            print("8. Generate Stock Chart")
            print("\nManagement:")
            print("9. Stock Adjustment")
            print("10. Undo Last Transaction")
            print("11. Export to Excel")
            print("12. Settings")
            print("0. Exit")
            choice = input("\nChoice: ").strip()
        
        try:
            if choice == '1':
                add_stock()
            elif choice == '2':
                record_sale()
            elif choice == '3':
                add_past_transaction_manual()
            elif choice == '4':
                bulk_entry_wizard()
            elif choice == '5':
                view_inventory()
            elif choice == '6':
                view_item_history()
            elif choice == '7':
                sales_and_profit_reports()
            elif choice == '8':
                generate_stock_chart()
            elif choice == '9':
                stock_adjustment()
            elif choice == '10':
                undo_last_transaction()
            elif choice == '11':
                export_to_excel()
            elif choice == '12':
                edit_settings()
            elif choice == '0':
                if RICH_AVAILABLE:
                    console.print("\n[green]💾 All data saved automatically![/green]")
                    console.print("[cyan]👋 Goodbye![/cyan]")
                else:
                    print("\n✓ Data saved. Goodbye!")
                break
            else:
                console.print("[red]Invalid choice.[/red]")
        
        except KeyboardInterrupt:
            console.print("\n\n[yellow]⚠️ Interrupted by user[/yellow]")
            if RICH_AVAILABLE:
                if Confirm.ask("Exit program?", default=False):
                    break
            else:
                break
        except Exception as e:
            log_error(f"Error in menu option {choice}: {str(e)}")
            console.print(f"\n[red]❌ Error: {str(e)}[/red]")
            console.print("[dim]Details logged to error_log.txt[/dim]")

def main():
    """Main entry point."""
    try:
        # Check if rich is available
        if not RICH_AVAILABLE:
            print("Installing required libraries...")
            os.system("pip install rich matplotlib")
            print("\nPlease restart the program.")
            return
        
        # Initialize database
        if not os.path.exists(DB_NAME):
            console.print("[yellow]Initializing new database...[/yellow]")
            init_database()
            console.print("[green]✓ Database created successfully![/green]")
        else:
            init_database()  # Ensure tables exist
        
        # Check for low stock on startup
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM transactions')
        count = cursor.fetchone()[0]
        conn.close()
        
        if count > 0:
            # Show low stock alerts
            dimensions = get_all_dimensions()
            low_stock_items = []
            
            for dim in dimensions:
                stock = get_current_stock(dim)
                if 0 < stock < CONFIG['low_stock_threshold']:
                    low_stock_items.append((dim, stock))
            
            if low_stock_items:
                console.print("\n[bold red]⚠️ LOW STOCK ALERTS:[/bold red]")
                for dim, stock in low_stock_items:
                    console.print(f"  [yellow]• {dim}: Only {stock:.2f} kg remaining[/yellow]")
        
        # Run main menu
        main_menu()
        
    except KeyboardInterrupt:
        console.print("\n\n[yellow]Program interrupted. Data is safe.[/yellow]")
    except Exception as e:
        log_error(f"Fatal error: {str(e)}")
        console.print(f"\n[bold red]❌ Fatal Error:[/bold red] {str(e)}")
        console.print("[dim]Details saved to error_log.txt[/dim]")

if __name__ == "__main__":
    main()